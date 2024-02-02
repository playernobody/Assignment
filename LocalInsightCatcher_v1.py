import googlemaps
import phonenumbers
import openpyxl
import requests
from openpyxl import load_workbook, Workbook
from phonenumbers import carrier

# Replace 'YOUR_API_KEY' with your Google API key
gmaps = googlemaps.Client(key='AIzaSyC3Mfkc0ib2yqkJ3cQCiJ2DDyTyeF8snno')

# Function to get GMB details
def get_gmb_details(place_id):
    try:
        result = gmaps.place(place_id=place_id, fields=[
            'name', 'formatted_address', 'international_phone_number', 'website', 'opening_hours', 'rating', 'reviews', 'geometry'])
        return result.get('result', {})
    except Exception as e:
        print(f"Error fetching details for place_id {place_id}: {e}")
        return {}

# Function to verify phone number on WhatsApp using Twilio
def is_whatsapp_number(phone_number):
    try:
        your_number = '+918580610536'
        check_number = phone_number
        url = "https://api.p.2chat.io/open/whatsapp/check-number/{}/{}?extra-information=false".format(your_number, check_number)
        payload = {}
        headers = {
            'X-User-API-Key': 'UAKbd8e26ad-01ab-4aaf-8859-9989a1185b2b'
        }
        response = requests.request("GET", url, headers=headers, data=payload)
        data = response.json()
        if data.get("on_whatsapp", True):
            return True
        else:
            return False
    except Exception as e:
        print(f"Error verifying WhatsApp for {phone_number}: {e}")
        return False

# Function to append data to Excel
def append_to_excel(data):
    try:
        try:
            # Load existing workbook if it exists
            wb = load_workbook("gmb_data.xlsx")
            ws = wb.active
        except FileNotFoundError:
            # Create a new workbook if the file doesn't exist
            wb = Workbook()
            ws = wb.active
            headers = ['Name', 'Address', 'Phone Number', 'Website', 'Opening Hours', 'Rating', 'Reviews', 'Latitude', 'Longitude', 'Is WhatsApp']
            ws.append(headers)

        for business in data:
            raw_phone_number = business.get('formatted_phone_number', '') or business.get('international_phone_number', '')
            cleaned_phone_number = ''.join(filter(str.isdigit, raw_phone_number))

            if cleaned_phone_number:
                try:
                    parsed_number = phonenumbers.parse("+" + cleaned_phone_number, None)
                    country_code = str(parsed_number.country_code)
                    national_number = str(parsed_number.national_number)
                    formatted_phone_number = f"+{country_code}{national_number}"
                except phonenumbers.NumberParseException as e:
                    print(f"Error parsing phone number {cleaned_phone_number}: {e}")
                    formatted_phone_number = cleaned_phone_number
            else:
                formatted_phone_number = ''

            row = [business.get('name', ''),
                   business.get('formatted_address', ''),
                   formatted_phone_number,
                   business.get('website', ''),
                   ', '.join(business.get('opening_hours', {}).get('weekday_text', [])),
                   business.get('rating', ''),
                   ', '.join([f"{review['author_name']}: {review['text']}" for review in business.get('reviews', [])]),
                   business.get('geometry', {}).get('location', {}).get('lat', ''),
                   business.get('geometry', {}).get('location', {}).get('lng', ''),
                   is_whatsapp_number(formatted_phone_number)
                   ]
            ws.append(row)

        wb.save("gmb_data.xlsx")
        print("Data appended to Excel.")
    except Exception as e:
        print(f"Error appending data to Excel: {e}")

# Replace 'YOUR_PLACE_ID' with the desired Google Place ID
# place_id = 'ChIJm-UHpDY_BTkR5jrKZyE1mSI'
place_id = 'ChIJFWFk-Ro_BTkRahriFH1FdkU'
gmb_data = get_gmb_details(place_id)

if gmb_data:
    append_to_excel([gmb_data])
